import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from tools.export_to_excel import ExcelExporter
from tools.generate_markdown_report import MarkdownReportGenerator
from tools.youtube_analyze_videos import YouTubeAnalyzer


def _video(
    video_id,
    title,
    duration,
    description,
    published_at,
    views,
    likes,
    comments,
    tags=None,
):
    return {
        "id": video_id,
        "title": title,
        "description": description,
        "publishedAt": published_at,
        "channelId": "UC_TEST",
        "channelTitle": "Test Channel",
        "tags": tags or ["test", "channel"],
        "categoryId": "22",
        "thumbnails": {},
        "duration": duration,
        "statistics": {
            "viewCount": views,
            "likeCount": likes,
            "commentCount": comments,
        },
    }


def _raw_data(videos):
    return {
        "channel": {
            "id": "UC_TEST",
            "title": "Test Channel",
            "description": "Test description",
            "customUrl": "",
            "publishedAt": "2020-01-01T00:00:00Z",
            "thumbnails": {},
            "subscriberCount": 1000,
            "videoCount": len(videos),
            "viewCount": sum(v["statistics"]["viewCount"] for v in videos),
        },
        "videos": videos,
        "metadata": {"quotaUsed": 10},
    }


class AnalyzerOutputs2026Tests(unittest.TestCase):
    def _mixed_videos(self):
        return [
            _video("v1", "Long Video One", "PT5M20S", "No chapters here", "2025-01-10T10:00:00Z", 5000, 300, 40),
            _video("v2", "Long Video Two", "PT8M00S", "0:00 Intro\n1:15 Main", "2025-01-12T10:00:00Z", 7000, 420, 55),
            _video("v3", "Short burst #shorts", "PT2M20S", "Fast clip #shorts", "2025-01-15T10:00:00Z", 12000, 900, 110),
            _video("v4", "Micro update", "PT45S", "Tiny update", "2025-01-16T10:00:00Z", 3000, 180, 8),
            _video("v5", "Another Long Deep Dive", "PT12M30S", "Detailed notes only", "2025-01-20T10:00:00Z", 9000, 600, 75),
            _video("v6", "Weekly recap", "PT3M10S", "Recap content", "2025-01-25T10:00:00Z", 2500, 140, 18),
        ]

    def test_shorts_detection_hybrid_rule(self):
        analyzer = YouTubeAnalyzer(_raw_data(self._mixed_videos()))

        self.assertTrue(analyzer.is_short_video(_video("s1", "x", "PT59S", "", "2025-01-01T00:00:00Z", 1, 0, 0)))
        self.assertTrue(analyzer.is_short_video(_video("s2", "topic #shorts", "PT2M30S", "", "2025-01-01T00:00:00Z", 1, 0, 0)))
        self.assertFalse(analyzer.is_short_video(_video("s3", "topic", "PT2M30S", "", "2025-01-01T00:00:00Z", 1, 0, 0)))
        self.assertFalse(analyzer.is_short_video(_video("s4", "#shorts", "PT3M1S", "", "2025-01-01T00:00:00Z", 1, 0, 0)))

    def test_timestamp_detection_patterns(self):
        analyzer = YouTubeAnalyzer(_raw_data(self._mixed_videos()))

        self.assertTrue(analyzer.is_timestamp_present("0:00 Intro\n1:02:03 walkthrough"))
        self.assertTrue(analyzer.is_timestamp_present("See section 12:34 for the demo"))
        self.assertFalse(analyzer.is_timestamp_present("No chapter markers listed here"))

    def test_timestamp_eligibility_excludes_shorts(self):
        videos = [
            _video("l1", "Long No TS", "PT6M00S", "No timestamps", "2025-01-01T00:00:00Z", 9000, 500, 60),
            _video("l2", "Long With TS", "PT4M00S", "0:00 intro", "2025-01-02T00:00:00Z", 8000, 480, 55),
            _video("s1", "Tagged #shorts", "PT2M30S", "clip #shorts", "2025-01-03T00:00:00Z", 12000, 700, 90),
            _video("s2", "Under 60", "PT40S", "quick short", "2025-01-04T00:00:00Z", 4000, 240, 12),
        ]
        analyzer = YouTubeAnalyzer(_raw_data(videos))
        shorts_videos, long_form_videos = analyzer.split_videos_by_format()
        timestamp_audit = analyzer.generate_timestamp_audit(long_form_videos)

        self.assertEqual(len(shorts_videos), 2)
        self.assertEqual(timestamp_audit["eligibleCount"], 2)
        self.assertEqual(timestamp_audit["missingCount"], 1)
        self.assertEqual(timestamp_audit["missingVideos"][0]["video_id"], "l1")

    def test_shorts_score_formula_and_floor(self):
        analyzer = YouTubeAnalyzer(_raw_data(self._mixed_videos()))
        score = analyzer.calculate_shorts_health_score(
            [{"priority": "High"}, {"priority": "Medium"}, {"priority": "Low"}]
        )
        self.assertEqual(score, 80)

        floor_score = analyzer.calculate_shorts_health_score([{"priority": "High"} for _ in range(20)])
        self.assertEqual(floor_score, 10)

    def test_generate_analysis_adds_new_schema_keys(self):
        raw = _raw_data(self._mixed_videos())
        analyzer = YouTubeAnalyzer(raw)
        analysis = analyzer.generate_analysis()

        self.assertIn("shortsHealthScore", analysis)
        self.assertIn("shortsRecommendations", analysis)
        self.assertIn("timestampAudit", analysis)
        self.assertIn("shorts2026", analysis.get("analysisModules", {}))

        summary = analysis.get("summary", {})
        self.assertIn("shortsVideos", summary)
        self.assertIn("longFormVideos", summary)
        self.assertIn("timestampEligibleVideos", summary)
        self.assertIn("timestampCoveragePercent", summary)
        self.assertIn("videoAudits", analysis.get("analysisModules", {}).get("shorts2026", {}))

    def test_excel_and_markdown_outputs_include_new_sections(self):
        raw = _raw_data(self._mixed_videos())
        analysis = YouTubeAnalyzer(raw).generate_analysis()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "audit.xlsx"
            ExcelExporter(raw, analysis).export(output_path)
            workbook = load_workbook(output_path)
            self.assertIn("Needs Timestamps", workbook.sheetnames)
            self.assertIn("Shorts Audit 2026", workbook.sheetnames)
            shorts_tab = workbook["Shorts Audit 2026"]
            values = [str(cell) for row in shorts_tab.iter_rows(values_only=True) for cell in row if cell is not None]
            joined = " ".join(values)
            self.assertIn("Shorts Videos and Optimization Opportunities", joined)
            self.assertIn("https://youtube.com/watch?v=", joined)

        markdown = MarkdownReportGenerator(raw, analysis).generate()
        self.assertIn("Timestamp Coverage Audit", markdown)
        self.assertIn("Shorts Audit (2026)", markdown)
        self.assertIn("Shorts Video Optimization Opportunities", markdown)
        self.assertIn("https://youtube.com/watch?v=", markdown)

    def test_no_shorts_channel_is_safe(self):
        videos = [
            _video("l1", "Long 1", "PT6M00S", "0:00 intro", "2025-01-01T00:00:00Z", 5000, 300, 40),
            _video("l2", "Long 2", "PT7M00S", "details", "2025-01-02T00:00:00Z", 4500, 250, 35),
            _video("l3", "Long 3", "PT8M00S", "details", "2025-01-03T00:00:00Z", 4300, 240, 32),
        ]
        analysis = YouTubeAnalyzer(_raw_data(videos)).generate_analysis()
        shorts = analysis.get("analysisModules", {}).get("shorts2026", {})

        self.assertEqual(shorts.get("shortsCount"), 0)
        self.assertEqual(analysis.get("shortsHealthScore"), 100)
        self.assertGreaterEqual(len(analysis.get("shortsRecommendations", [])), 1)
        self.assertEqual(shorts.get("videoAudits"), [])

    def test_all_shorts_channel_has_zero_timestamp_eligibility(self):
        videos = [
            _video("s1", "A #shorts", "PT35S", "clip", "2025-01-01T00:00:00Z", 1500, 80, 9),
            _video("s2", "B #shorts", "PT2M20S", "clip #shorts", "2025-01-02T00:00:00Z", 2100, 120, 14),
            _video("s3", "C #shorts", "PT55S", "clip", "2025-01-03T00:00:00Z", 1900, 95, 10),
        ]
        analysis = YouTubeAnalyzer(_raw_data(videos)).generate_analysis()
        self.assertEqual(analysis.get("summary", {}).get("longFormVideos"), 0)
        self.assertEqual(analysis.get("timestampAudit", {}).get("eligibleCount"), 0)


if __name__ == "__main__":
    unittest.main()
