#!/usr/bin/env python3
"""
Excel Exporter
Creates a multi-tab Excel workbook from audit data.

Usage:
    python3 export_to_excel.py path/to/raw_data.json path/to/analysis.json [output.xlsx]
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")


def has_timestamps(description):
    return bool(description) and bool(re.search(r"(?<!\d)(?:\d{1,2}:\d{2}(?::\d{2})?)(?!\d)", description))


def autosize_columns(worksheet, max_width=80):
    """Auto-size columns to content width with a reasonable cap."""
    widths = {}
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            widths[cell.column] = max(widths.get(cell.column, 0), length)

    for col_idx, width in widths.items():
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), max_width)


def style_title_row(worksheet, end_column):
    """Style and merge row 1 as title."""
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = worksheet.cell(row=1, column=1)
    cell.fill = TITLE_FILL
    cell.font = Font(bold=True, color="FFFFFF", size=13)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header_row(worksheet, row_idx, end_column):
    """Style a header row."""
    for col in range(1, end_column + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_section_row(worksheet, row_idx, end_column):
    """Style section rows for readability."""
    for col in range(1, end_column + 1):
        cell = worksheet.cell(row=row_idx, column=col)
        cell.fill = SECTION_FILL
        cell.font = Font(bold=True)


class ExcelExporter:
    def __init__(self, raw_data, analysis):
        self.raw_data = raw_data
        self.analysis = analysis
        self.channel = raw_data["channel"]
        self.videos = raw_data["videos"]

    def create_summary_tab(self, workbook):
        ws = workbook.create_sheet("Summary")
        health_score = self.analysis.get("channelHealthScore", 0)
        shorts_health_score = self.analysis.get("shortsHealthScore")
        summary = self.analysis.get("summary", {})
        timestamp_coverage = summary.get("timestampCoveragePercent", 0.0)

        rows = [
            ["YOUTUBE CHANNEL AUDIT - EXECUTIVE SUMMARY"],
            [""],
            ["Channel Information", ""],
            ["Channel Name", self.channel.get("title", "")],
            ["Subscribers", self.channel.get("subscriberCount", 0)],
            ["Total Videos", self.channel.get("videoCount", 0)],
            ["Total Views", self.channel.get("viewCount", 0)],
            [""],
            ["Audit Results", ""],
            ["Channel Health Score", f"{health_score}/100"],
            ["Videos Analyzed", len(self.videos)],
            ["Total Recommendations", summary.get("totalRecommendations", 0)],
            ["High Priority Issues", summary.get("highPriority", 0)],
            ["Medium Priority Issues", summary.get("mediumPriority", 0)],
            ["Low Priority Issues", summary.get("lowPriority", 0)],
            ["Shorts Health Score", f"{shorts_health_score}/100" if shorts_health_score is not None else "N/A"],
            ["Shorts Videos", summary.get("shortsVideos", 0)],
            ["Long-form Videos", summary.get("longFormVideos", 0)],
            ["Timestamp Coverage", f"{timestamp_coverage}%"],
            ["Videos Missing Timestamps", summary.get("timestampMissingVideos", 0)],
            [""],
            ["Top 5 Recommendations", ""],
        ]

        for idx, rec in enumerate(self.analysis.get("allRecommendations", [])[:5], 1):
            rows.append([f"{idx}. [{rec.get('priority', 'N/A')}] {rec.get('category', 'N/A')}", rec.get("recommendation", "")])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 2)
        style_section_row(ws, 3, 2)
        style_section_row(ws, 9, 2)
        style_section_row(ws, 22, 2)
        ws.freeze_panes = "A4"
        autosize_columns(ws)

    def create_scoring_methodology_tab(self, workbook):
        ws = workbook.create_sheet("Scoring Methodology")
        health_score = self.analysis.get("channelHealthScore", 0)
        shorts_health_score = self.analysis.get("shortsHealthScore")
        summary = self.analysis.get("summary", {})

        rows = [
            ["SCORING METHODOLOGY"],
            [""],
            ["Channel Health Score", "", "", "", ""],
            ["Formula", "max(10, 100 - (10 x High Priority Issues) - (5 x Medium Priority Issues))", "", "", ""],
            ["Your Score", f"{health_score}/100", "", "", ""],
            ["High Priority Issues", summary.get("highPriority", 0), "", "", ""],
            ["Medium Priority Issues", summary.get("mediumPriority", 0), "", "", ""],
            ["Low Priority Issues", summary.get("lowPriority", 0), "(not penalized in channel score)", "", ""],
            [""],
            ["Score Interpretation", "", "", "", ""],
            ["Range", "Rating", "Meaning", "Action", ""],
            ["80-100", "Excellent", "Minor improvements needed", "Focus on low-priority optimizations", ""],
            ["60-79", "Good", "Solid baseline with gaps", "Address medium-priority items", ""],
            ["40-59", "Needs Work", "Significant optimization gaps", "Prioritize high-priority issues", ""],
            ["20-39", "Critical", "Major discoverability risk", "Urgent fixes across key areas", ""],
            ["10-19", "Severe", "Extensive channel issues", "Full optimization pass", ""],
            [""],
            ["Shorts Health Score", "", "", "", ""],
            ["Formula", "max(10, 100 - (12 x High) - (6 x Medium) - (2 x Low))", "", "", ""],
            ["Your Shorts Score", f"{shorts_health_score}/100" if shorts_health_score is not None else "N/A", "", "", ""],
            ["Note", "Shorts use more aggressive penalties and include Low priority deductions", "", "", ""],
            [""],
            ["Industry Benchmarks (Long-Form)", "", "", "", ""],
            ["Metric", "Benchmark", "Why it matters", "Calculation", "Source"],
            ["Title Length", "40-70 characters (optimal)", "Long enough for keywords, short enough to avoid truncation", "Character count", "YouTube Creator Academy & TubeBuddy 2024 Research"],
            ["Description Length", "100 min, 300+ optimal", "First 150 chars appear in search; longer descriptions improve algorithm context", "Character count", "YouTube SEO Best Practices 2024"],
            ["Tags Per Video", "8-12 tags (optimal)", "Balances content context coverage and relevance", "Tag count", "YouTube Metadata Optimization Guidelines"],
            ["Engagement Rate", "<2% poor \u00b7 4%+ good \u00b7 6%+ excellent", "Signals content quality to algorithm, improves search rankings", "((likes+comments)/views)\u00d7100", "Social Blade & VidIQ Analytics 2024"],
            ["Comments per 1K", "5 min \u00b7 10 good \u00b7 20+ excellent", "Comments heavily weighted by algorithm for search and suggestions", "(comments/views)\u00d71000", "YouTube Algorithm Research 2024"],
            ["Upload Frequency", "1-2/week optimal", "Consistent uploads train audience and signal active channel", "Uploads per week", "YouTube Creator Insider 2024"],
            [""],
            ["Shorts-Specific Benchmarks", "", "", "", ""],
            ["Metric", "Benchmark", "Why it matters", "Threshold", "Source"],
            ["Title Length", "20-70 characters", "Concise, intent-led titles improve tap propensity", "\u226560% of Shorts in range", "YouTube Creator Academy (official)"],
            ["Description", "\u226540 chars or include hashtags", "Basic context aids categorization and discovery", "\u226450% sparse without hashtags", "YouTube Help Center + Creator Academy (official)"],
            ["Posting Freshness", "\u22651 Short every 21 days", "Fresh cadence supports Shorts distribution momentum", "Active channels only", "YouTube Creator Insider (official)"],
            ["Engagement (vs Long-form)", "\u226580% of long-form median", "Relative normalization avoids false positives on niche channels", "Channels with \u226510 Shorts", "Conservative heuristic (official-source-aligned)"],
            ["Comments / 1K Views", "\u22655 per 1K", "Comment activity signals resonance beyond passive views", "Channel aggregate", "YouTube community interaction guidance"],
        ]

        for row in rows:
            ws.append(row)

        style_title_row(ws, 5)
        style_section_row(ws, 3, 5)
        style_section_row(ws, 10, 5)
        style_section_row(ws, 18, 5)
        style_section_row(ws, 23, 5)
        style_section_row(ws, 32, 5)
        style_header_row(ws, 11, 5)
        style_header_row(ws, 24, 5)
        style_header_row(ws, 33, 5)
        ws.freeze_panes = "A4"
        autosize_columns(ws, max_width=70)

    def create_audit_checklist_tab(self, workbook):
        ws = workbook.create_sheet("Audit Checklist")
        checklist = self.analysis.get("auditChecklist", {})
        summary = checklist.get("summary", {})

        rows = [
            ["YOUTUBE CHANNEL AUDIT - DIAGNOSTIC CHECKLIST"],
            [""],
            ["Summary", "", "", "", ""],
            ["Total Issues Found", summary.get("total_issues", 0), "", "", ""],
            ["Critical Issues", summary.get("critical_issues", 0), "", "", ""],
            ["Warnings", summary.get("warnings", 0), "", "", ""],
            ["Quick Wins Available", summary.get("quick_wins", 0), "", "", ""],
            ["Estimated Fix Time", summary.get("estimated_fix_time", "N/A"), "", "", ""],
            ["Potential Impact", summary.get("potential_impact", "N/A"), "", "", ""],
            [""],
            ["Critical Issues (Must Fix)", "", "", "", ""],
            ["Issue Type", "Count", "% of Videos", "Impact", "Severity"],
        ]

        for issue in checklist.get("critical_issues", []):
            rows.append([
                issue.get("issue", ""),
                issue.get("count", 0),
                issue.get("percentage", ""),
                issue.get("impact", ""),
                issue.get("severity", ""),
            ])

        rows.extend([
            [""],
            ["Engagement Warnings", "", "", "", ""],
            ["Metric", "Your Channel", "Benchmark", "Gap", "Status"],
        ])

        for warning in checklist.get("engagement_warnings", []):
            rows.append([
                warning.get("issue", ""),
                warning.get("current", ""),
                warning.get("benchmark", ""),
                warning.get("gap", ""),
                warning.get("status", ""),
            ])

        rows.extend([
            [""],
            ["Upload Schedule Issues", "", "", "", ""],
            ["Issue Type", "Current", "Benchmark", "Status", ""],
        ])

        for item in checklist.get("upload_schedule_issues", []):
            rows.append([
                item.get("issue", ""),
                item.get("current", ""),
                item.get("benchmark", ""),
                item.get("status", ""),
                "",
            ])

        rows.extend([
            [""],
            ["Optimization Opportunities", "", "", "", ""],
            ["Issue Type", "Count", "Quick Fix?", "Expected Impact", ""],
        ])

        for item in checklist.get("optimization_opportunities", []):
            rows.append([
                item.get("issue", ""),
                item.get("count", 0),
                item.get("quick_fix", ""),
                item.get("impact", ""),
                "",
            ])

        shorts_recs = self.analysis.get("shortsRecommendations", [])
        if shorts_recs:
            rows.extend([
                [""],
                ["Shorts Issues", "", "", "", ""],
                ["Issue", "Priority", "Benchmark", "Recommendation", "Impact"],
            ])
            for rec in shorts_recs:
                rows.append([
                    rec.get("issue", ""),
                    rec.get("priority", ""),
                    rec.get("benchmark", "N/A"),
                    rec.get("recommendation", ""),
                    rec.get("impact", ""),
                ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 5)
        style_section_row(ws, 3, 5)

        for row_num in range(1, ws.max_row + 1):
            first_value = ws.cell(row=row_num, column=1).value
            if isinstance(first_value, str) and first_value in {
                "Critical Issues (Must Fix)",
                "Engagement Warnings",
                "Upload Schedule Issues",
                "Optimization Opportunities",
                "Shorts Issues",
            }:
                style_section_row(ws, row_num, 5)

        for row_num in range(1, ws.max_row + 1):
            first_value = ws.cell(row=row_num, column=1).value
            if first_value in {"Issue Type", "Metric", "Issue"}:
                style_header_row(ws, row_num, 5)

        ws.freeze_panes = "A4"
        autosize_columns(ws, max_width=65)

    def create_quick_wins_tab(self, workbook):
        ws = workbook.create_sheet("Quick Wins")
        rows = [
            ["QUICK WINS - IMMEDIATE ACTION ITEMS"],
            [""],
            ["Priority", "Action Type", "Video URL", "Video Title", "Current State", "Suggested Fix", "Expected Impact", "Effort"],
        ]

        for qw in self.analysis.get("quickWins", []):
            rows.append([
                qw.get("priority", ""),
                qw.get("action", ""),
                qw.get("video_url", ""),
                qw.get("video_title", ""),
                qw.get("current_state", ""),
                qw.get("suggested_fix", ""),
                qw.get("impact", ""),
                qw.get("effort", ""),
            ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 8)
        style_header_row(ws, 3, 8)
        ws.freeze_panes = "A4"
        autosize_columns(ws, max_width=65)

    def create_before_after_tab(self, workbook):
        ws = workbook.create_sheet("Before After")
        rows = [
            ["BEFORE/AFTER OPTIMIZATION EXAMPLES"],
            [""],
            ["Type", "Video URL", "Before", "After", "Why It Is Better", "Expected Impact"],
        ]

        for item in self.analysis.get("beforeAfterExamples", []):
            rows.append([
                item.get("type", ""),
                item.get("video_url", ""),
                item.get("before", ""),
                item.get("after", ""),
                item.get("why_better", ""),
                item.get("impact", ""),
            ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 6)
        style_header_row(ws, 3, 6)
        ws.freeze_panes = "A4"
        autosize_columns(ws, max_width=70)

    def create_video_performance_tab(self, workbook):
        ws = workbook.create_sheet("Video Performance")
        headers = [
            "Video URL",
            "Title",
            "Views",
            "Likes",
            "Comments",
            "Engagement Rate",
            "Published Date",
            "Tags Count",
            "Title Length",
            "Description Length",
            "Performance Tier",
        ]
        ws.append(headers)

        sorted_videos = sorted(self.videos, key=lambda video: video["statistics"]["viewCount"], reverse=True)
        total = len(sorted_videos)
        high_cutoff = max(1, total // 3)
        medium_cutoff = max(1, (2 * total) // 3)

        for idx, video in enumerate(sorted_videos):
            views = video["statistics"]["viewCount"]
            likes = video["statistics"]["likeCount"]
            comments = video["statistics"]["commentCount"]
            engagement_rate = ((likes + comments) / views * 100) if views else 0.0

            if idx < high_cutoff:
                tier = "High"
            elif idx < medium_cutoff:
                tier = "Medium"
            else:
                tier = "Low"

            ws.append([
                f"https://youtube.com/watch?v={video['id']}",
                video["title"],
                views,
                likes,
                comments,
                round(engagement_rate, 2),
                video["publishedAt"][:10],
                len(video.get("tags", [])),
                len(video["title"]),
                len(video.get("description", "")),
                tier,
            ])

        style_header_row(ws, 1, 11)
        ws.freeze_panes = "A2"
        autosize_columns(ws, max_width=70)

    def create_title_description_tab(self, workbook):
        ws = workbook.create_sheet("Title Description Audit")
        titles = self.analysis.get("analysisModules", {}).get("titlesAndDescriptions", {})

        rows = [
            ["TITLE & DESCRIPTION ANALYSIS"],
            [""],
            ["Key Metrics", "", "", "", "", "", ""],
            ["Average Title Length", f"{titles.get('titleLengthAverage', 0)} characters", "", "", "", "", ""],
            ["High Performers Avg Title", f"{titles.get('titleLengthHighPerformers', 0)} characters", "", "", "", "", ""],
            ["Average Description Length", f"{titles.get('descriptionLengthAverage', 0)} characters", "", "", "", "", ""],
            ["Videos with Timestamps", titles.get("videosWithTimestamps", 0), "", "", "", "", ""],
            [""],
            ["Common Keywords in Top Performers", "", "", "", "", "", ""],
        ]

        for keyword, count in titles.get("commonKeywords", []):
            rows.append([keyword, count, "", "", "", "", ""])

        rows.extend([
            [""],
            ["Video-by-Video Analysis", "", "", "", "", "", ""],
            ["Video URL", "Title", "Title Length", "Description Length", "Has Timestamps", "Views", "Status"],
        ])

        for video in self.videos:
            title_len = len(video["title"])
            description = video.get("description", "")
            desc_len = len(description)
            contains_timestamps = "Yes" if has_timestamps(description) else "No"

            status = "Good"
            if title_len > 70 or title_len < 40:
                status = "Review Title Length"
            if desc_len < 100:
                status = "Expand Description"

            rows.append([
                f"https://youtube.com/watch?v={video['id']}",
                video["title"],
                title_len,
                desc_len,
                contains_timestamps,
                video["statistics"]["viewCount"],
                status,
            ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 7)
        style_section_row(ws, 3, 7)
        style_section_row(ws, 9, 7)

        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == "Video URL":
                style_header_row(ws, row_num, 7)
                break

        ws.freeze_panes = "A14"
        autosize_columns(ws, max_width=70)

    def create_needs_timestamps_tab(self, workbook):
        ws = workbook.create_sheet("Needs Timestamps")
        timestamp_audit = self.analysis.get("timestampAudit", {})
        missing_videos = timestamp_audit.get("missingVideos", [])
        eligible_count = timestamp_audit.get("eligibleCount", 0)
        with_timestamps_count = timestamp_audit.get("withTimestampsCount", 0)
        missing_count = timestamp_audit.get("missingCount", 0)
        coverage_percent = timestamp_audit.get("coveragePercent", 0.0)

        rows = [
            ["VIDEOS NEEDING TIMESTAMPS (>2 MIN LONG-FORM)"],
            [""],
            ["Metric", "Value", "", "", "", "", "", ""],
            ["Eligible Videos (>2 min)", eligible_count, "", "", "", "", "", ""],
            ["With Timestamps", with_timestamps_count, "", "", "", "", "", ""],
            ["Missing Timestamps", missing_count, "", "", "", "", "", ""],
            ["Coverage %", f"{coverage_percent}%", "", "", "", "", "", ""],
            [""],
            ["Priority", "Video URL", "Title", "Duration (min)", "Views", "Published", "Description Length", "Has Timestamps"],
        ]

        if eligible_count == 0:
            rows.append(["Info", "N/A", "No eligible videos over 2 minutes.", "", "", "", "", ""])
        elif not missing_videos:
            rows.append(["Info", "N/A", "All eligible videos already have timestamps.", "", "", "", "", ""])
        else:
            for item in missing_videos:
                rows.append([
                    item.get("priority", "Medium"),
                    item.get("video_url", ""),
                    item.get("title", ""),
                    item.get("duration_minutes", ""),
                    item.get("views", 0),
                    str(item.get("publishedAt", ""))[:10],
                    item.get("description_length", 0),
                    item.get("has_timestamps", "No"),
                ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 8)
        style_header_row(ws, 3, 8)
        style_header_row(ws, 9, 8)
        ws.freeze_panes = "A10"
        autosize_columns(ws, max_width=70)

    def create_shorts_audit_tab(self, workbook):
        ws = workbook.create_sheet("Shorts Audit 2026")
        shorts_analysis = self.analysis.get("analysisModules", {}).get("shorts2026", {})
        shorts_recommendations = self.analysis.get("shortsRecommendations", shorts_analysis.get("recommendations", []))
        shorts_health_score = self.analysis.get("shortsHealthScore")
        shorts_video_audits = shorts_analysis.get("videoAudits", [])

        rows = [
            ["SHORTS AUDIT (2026)"],
            [""],
            ["Metric", "Value", "", "", "", ""],
            ["Shorts Health Score", f"{shorts_health_score}/100" if shorts_health_score is not None else "N/A", "", "", "", ""],
            ["Shorts Count", shorts_analysis.get("shortsCount", 0), "", "", "", ""],
            ["Shorts % of Channel", f"{shorts_analysis.get('shortsPercentOfChannel', 0)}%", "", "", "", ""],
            ["Avg Duration (sec)", shorts_analysis.get("avgDurationSeconds", 0), "", "", "", ""],
            ["Avg Views", shorts_analysis.get("avgViews", 0), "", "", "", ""],
            ["Median Views", shorts_analysis.get("medianViews", 0), "", "", "", ""],
            ["Avg Engagement", f"{shorts_analysis.get('avgEngagementRate', 0)}%", "", "", "", ""],
            ["Comments / 1K Views", shorts_analysis.get("commentsPer1k", 0), "", "", "", ""],
            ["Last 30d Shorts", shorts_analysis.get("postedLast30Days", 0), "", "", "", ""],
            ["Days Since Last Short", shorts_analysis.get("daysSinceLastShort", "N/A"), "", "", "", ""],
            ["Metadata Coverage", f"{shorts_analysis.get('metadataCoverage', 0)}%", "", "", "", ""],
            ["Videos With Opportunities", shorts_analysis.get("videosWithOpportunities", 0), "", "", "", ""],
            ["Total Video Opportunities", shorts_analysis.get("totalVideoOpportunities", 0), "", "", "", ""],
            [""],
            ["Recommendations", "", "", "", "", ""],
            ["Priority", "Issue", "Benchmark", "Recommendation", "Impact", "Source"],
        ]

        if shorts_analysis.get("shortsCount", 0) == 0:
            rows.append([
                "Informational",
                "No Shorts identified by configured rule.",
                "N/A",
                "If Shorts are in scope, publish pilots and re-audit.",
                "Establishes Shorts baseline",
                "YouTube Help + Creator Academy",
            ])
        elif not shorts_recommendations:
            rows.append([
                "Info",
                "No Shorts-specific issues triggered by configured checks.",
                "N/A",
                "Maintain current Shorts approach and monitor.",
                "Sustains Shorts performance",
                "Audit rule output",
            ])
        else:
            for rec in shorts_recommendations:
                rows.append([
                    rec.get("priority", "Low"),
                    rec.get("issue", ""),
                    rec.get("benchmark", "N/A"),
                    rec.get("recommendation", ""),
                    rec.get("impact", ""),
                    rec.get("source", ""),
                ])

        rows.extend([
            [""],
            ["Shorts Video Audit", "", "", "", "", "", ""],
            ["Video URL", "Title", "Views", "Engagement %", "Comments / 1K", "Published", "Audit Recommendation"],
        ])

        if not shorts_video_audits:
            rows.append([
                "N/A",
                "No Shorts identified by configured rule.",
                "",
                "",
                "",
                "",
                "",
            ])
        else:
            for item in shorts_video_audits:
                opportunities = item.get("optimizationOpportunities", [])
                base = [
                    item.get("video_url", ""),
                    item.get("title", ""),
                    item.get("views", 0),
                    f"{item.get('engagementRate', 0)}%",
                    item.get("commentsPer1k", 0),
                    item.get("publishedAt", "")[:10],
                ]
                no_issues = ["No immediate opportunities flagged by configured checks."]
                if not opportunities or opportunities == no_issues:
                    rows.append(base + ["No issues flagged"])
                else:
                    for opp in opportunities:
                        rows.append(base + [opp])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 7)
        style_header_row(ws, 3, 7)

        for row_num in range(1, ws.max_row + 1):
            first_value = ws.cell(row=row_num, column=1).value
            if first_value in {"Recommendations", "Shorts Video Audit"}:
                style_section_row(ws, row_num, 7)
            if first_value in {"Priority", "Video URL"}:
                style_header_row(ws, row_num, 7)

        ws.freeze_panes = "A18"
        autosize_columns(ws, max_width=72)

    def create_tags_metadata_tab(self, workbook):
        ws = workbook.create_sheet("Tags Metadata")
        tags = self.analysis["analysisModules"]["tagsAndMetadata"]

        rows = [
            ["TAGS & METADATA ANALYSIS"],
            [""],
            ["Key Metrics", "", "", "", ""],
            ["Average Tags Per Video", tags["averageTagCount"], "", "", ""],
            ["Videos Without Tags", tags["videosWithoutTags"], "", "", ""],
            ["Category Consistency", f"{tags['categoryConsistency']}%", "", "", ""],
            ["Most Common Category", tags["mostCommonCategory"], "", "", ""],
            [""],
            ["Most Common Tags", "", "", "", ""],
            ["Tag", "Count", "Coverage", "", ""],
        ]

        for tag, count in tags.get("commonTags", []):
            coverage = (count / len(self.videos) * 100) if self.videos else 0
            rows.append([tag, count, f"{coverage:.1f}%", "", ""])

        rows.extend([
            [""],
            ["Video-by-Video Tag Analysis", "", "", "", ""],
            ["Video URL", "Title", "Tag Count", "Tags Preview", "Views"],
        ])

        for video in self.videos:
            tag_values = video.get("tags", [])
            rows.append([
                f"https://youtube.com/watch?v={video['id']}",
                video["title"],
                len(tag_values),
                ", ".join(tag_values[:5]) if tag_values else "(no tags)",
                video["statistics"]["viewCount"],
            ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 5)
        style_section_row(ws, 3, 5)
        style_section_row(ws, 9, 5)
        style_header_row(ws, 10, 5)

        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == "Video URL":
                style_header_row(ws, row_num, 5)
                break

        ws.freeze_panes = "A13"
        autosize_columns(ws, max_width=70)

    def create_engagement_tab(self, workbook):
        ws = workbook.create_sheet("Engagement Analysis")
        engagement = self.analysis["analysisModules"]["engagement"]

        rows = [
            ["ENGAGEMENT ANALYSIS"],
            [""],
            ["Key Metrics", "", ""],
            ["Average Engagement Rate", f"{engagement['averageEngagementRate']}%", ""],
            ["Likes per 1000 Views", engagement["likesPerThousandViews"], ""],
            ["Comments per 1000 Views", engagement["commentsPerThousandViews"], ""],
            [""],
            ["Top 5 Most Engaging Videos", "", ""],
            ["Title", "Engagement Rate", "Views"],
        ]

        for video in engagement.get("topPerformers", []):
            rows.append([video.get("title", ""), f"{video.get('engagementRate', 0)}%", video.get("views", 0)])

        rows.extend([
            [""],
            ["Bottom 5 Least Engaging Videos", "", ""],
            ["Title", "Engagement Rate", "Views"],
        ])

        for video in engagement.get("bottomPerformers", []):
            rows.append([video.get("title", ""), f"{video.get('engagementRate', 0)}%", video.get("views", 0)])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 3)
        style_section_row(ws, 3, 3)
        style_section_row(ws, 8, 3)
        style_header_row(ws, 9, 3)

        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == "Bottom 5 Least Engaging Videos":
                style_section_row(ws, row_num, 3)
                style_header_row(ws, row_num + 1, 3)
                break

        ws.freeze_panes = "A10"
        autosize_columns(ws, max_width=70)

    def create_upload_schedule_tab(self, workbook):
        ws = workbook.create_sheet("Upload Schedule")
        schedule = self.analysis["analysisModules"]["uploadSchedule"]

        if "error" in schedule:
            rows = [
                ["UPLOAD SCHEDULE ANALYSIS"],
                [""],
                ["Error", schedule["error"], "", ""],
            ]
            for row in rows:
                ws.append(row)
            style_title_row(ws, 4)
            autosize_columns(ws)
            return

        rows = [
            ["UPLOAD SCHEDULE ANALYSIS"],
            [""],
            ["Key Metrics", "", "", ""],
            ["Average Gap Between Uploads", f"{schedule['averageGapDays']} days", "", ""],
            ["Consistency Score", f"{schedule['consistencyScore']}/10", "", ""],
            ["Uploads Per Week", schedule["uploadsPerWeek"], "", ""],
            ["Days Since Last Upload", schedule["daysSinceLastUpload"], "", ""],
            [""],
            ["Best Performing Days", "", "", ""],
            ["Day of Week", "Avg Views", "Upload Count", ""],
        ]

        for day, avg_views in schedule.get("bestPerformingDays", []):
            count = schedule.get("uploadDistribution", {}).get(day, {}).get("count", 0)
            rows.append([day, int(avg_views), count, ""])

        rows.extend([
            [""],
            ["Upload Distribution by Day", "", "", ""],
            ["Day", "Uploads", "Average Views", ""],
        ])

        for day, values in schedule.get("uploadDistribution", {}).items():
            rows.append([day, values.get("count", 0), int(values.get("views", 0)), ""])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 4)
        style_section_row(ws, 3, 4)
        style_section_row(ws, 9, 4)
        style_header_row(ws, 10, 4)

        for row_num in range(1, ws.max_row + 1):
            if ws.cell(row=row_num, column=1).value == "Upload Distribution by Day":
                style_section_row(ws, row_num, 4)
                style_header_row(ws, row_num + 1, 4)
                break

        ws.freeze_panes = "A11"
        autosize_columns(ws, max_width=60)

    def create_action_items_tab(self, workbook):
        ws = workbook.create_sheet("Action Items")
        rows = [
            ["ACTION ITEMS - PRIORITIZED RECOMMENDATIONS"],
            [""],
            ["Priority", "Category", "Issue", "Industry Benchmark", "Why This Matters", "Recommendation", "Expected Impact"],
        ]

        combined_recommendations = list(self.analysis.get("allRecommendations", []))
        combined_recommendations.extend(self.analysis.get("shortsRecommendations", []))
        priority_order = {"High": 0, "Medium": 1, "Low": 2, "Informational": 3, "Info": 4}
        combined_recommendations.sort(
            key=lambda rec: (
                priority_order.get(rec.get("priority", "Low"), 5),
                rec.get("category", ""),
            )
        )

        for rec in combined_recommendations:
            rows.append([
                rec.get("priority", ""),
                rec.get("category", ""),
                rec.get("issue", ""),
                rec.get("benchmark", "N/A"),
                rec.get("why", ""),
                rec.get("recommendation", ""),
                rec.get("impact", ""),
            ])

        for row in rows:
            ws.append(row)

        style_title_row(ws, 7)
        style_header_row(ws, 3, 7)
        ws.freeze_panes = "A4"
        autosize_columns(ws, max_width=70)

    def export(self, output_path):
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self.create_summary_tab(workbook)
        self.create_scoring_methodology_tab(workbook)
        self.create_audit_checklist_tab(workbook)
        self.create_quick_wins_tab(workbook)
        self.create_before_after_tab(workbook)
        self.create_video_performance_tab(workbook)
        self.create_title_description_tab(workbook)
        self.create_needs_timestamps_tab(workbook)
        self.create_shorts_audit_tab(workbook)
        self.create_tags_metadata_tab(workbook)
        self.create_engagement_tab(workbook)
        self.create_upload_schedule_tab(workbook)
        self.create_action_items_tab(workbook)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path


def main():
    if len(sys.argv) < 3 or len(sys.argv) > 4:
        print("Error: Missing required files")
        print("\nUsage:")
        print("  python3 export_to_excel.py path/to/raw_data.json path/to/analysis.json [output.xlsx]")
        sys.exit(1)

    raw_data_file = Path(sys.argv[1])
    analysis_file = Path(sys.argv[2])
    output_file = Path(sys.argv[3]) if len(sys.argv) == 4 else raw_data_file.parent / "audit_report.xlsx"

    try:
        print("Loading data files...")
        with raw_data_file.open("r", encoding="utf-8") as f:
            raw_data = json.load(f)
        with analysis_file.open("r", encoding="utf-8") as f:
            analysis = json.load(f)

        print("Exporting to Excel workbook...")
        print("=" * 50)

        exporter = ExcelExporter(raw_data, analysis)
        saved_path = exporter.export(output_file)

        print("\n" + "=" * 50)
        print("SUCCESS")
        print(f"\nExcel file saved at:\n{saved_path}")
        print(f"\nGenerated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    except FileNotFoundError as exc:
        print(f"Error: File not found: {exc}")
        sys.exit(1)
    except json.JSONDecodeError as exc:
        print(f"Error: Invalid JSON file: {exc}")
        sys.exit(1)
    except Exception as exc:
        print(f"Error: {exc}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
